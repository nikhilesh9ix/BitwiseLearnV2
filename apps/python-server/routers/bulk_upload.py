from datetime import datetime, timezone
from fastapi import APIRouter, Depends, UploadFile, File
from beanie import PydanticObjectId
from utils.api_response import api_response
from utils.password import hash_password
from middleware.auth import get_current_user, not_student, admin_only
from models.student import Student
from models.batch import Batch
from models.institution import Institution
from models.problem_test_case import ProblemTestCase
from models.assessment import Assessment
from models.assessment_section import AssessmentSection
from models.assessment_question import AssessmentQuestion
from enums import TestcaseType, AssessmentType
import openpyxl
import io
import secrets
import string
import re
import csv

router = APIRouter(prefix="/api/v1/bulk-upload", tags=["Bulk Upload"])


def _generate_password(length: int = 10) -> str:
    chars = string.ascii_letters + string.digits
    return "".join(secrets.choice(chars) for _ in range(length))


@router.post("/students/{id}")
async def bulk_upload_students(
    id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(not_student),
):
    batch_id = id
    # Try to find the batch to get institution_id
    batch = await Batch.get(PydanticObjectId(batch_id))
    if not batch:
        return api_response(404, "Batch not found", error="Not found")
    institution_id = str(batch.institution_id)

    contents = await file.read()

    rows: list[tuple] = []
    file_name = (file.filename or "").lower()
    if file_name.endswith(".csv"):
        text = contents.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = [tuple(row) for row in reader]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

    created = 0
    errors: list[str] = []

    def _looks_like_email(value: str) -> bool:
        return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value or ""))

    for idx, row in enumerate(rows[1:], start=2):
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()

        # Supported formats:
        # 1) New: [name, roll_number, email, batch_name]
        # 2) Legacy: [name, email, phone]
        col2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        col3 = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        col4 = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if _looks_like_email(col2):
            # Legacy sheet without roll number
            email = col2
            roll_number = f"AUTO-{idx}"
            batch_name = ""
        else:
            roll_number = col2
            email = col3
            batch_name = col4

        if not roll_number:
            errors.append(f"Row {idx}: Missing roll number")
            continue

        if not email:
            errors.append(f"Row {idx}: Missing email")
            continue

        if not _looks_like_email(email):
            errors.append(f"Row {idx}: Invalid email '{email}'")
            continue

        if batch_name and batch_name.strip().lower() != str(batch.batchname).strip().lower():
            errors.append(
                f"Row {idx}: Batch name '{batch_name}' does not match current batch '{batch.batchname}'"
            )
            continue

        existing = await Student.find_one(Student.email == email)
        if existing:
            errors.append(f"Row {idx}: Email {email} already exists")
            continue

        existing_roll = await Student.find_one(
            Student.roll_number == roll_number,
            Student.batch_id == PydanticObjectId(batch_id),
        )
        if existing_roll:
            errors.append(f"Row {idx}: Roll number {roll_number} already exists in this batch")
            continue

        password = _generate_password()
        student = Student(
            name=name,
            roll_number=roll_number,
            email=email,
            login_password=hash_password(password),
            institute_id=PydanticObjectId(institution_id),
            batch_id=PydanticObjectId(batch_id),
        )
        await student.insert()
        created += 1

    return api_response(200, "Students uploaded", data={"created": created, "errors": errors})


@router.post("/batches/{id}")
async def bulk_upload_batches(
    id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(not_student),
):
    institution_id = id

    institution = await Institution.get(PydanticObjectId(institution_id))
    if not institution:
        return api_response(404, "Institution not found", error="Not found")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    created = 0
    errors: list[str] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()

        existing = await Batch.find_one(Batch.name == name, Batch.institution_id == PydanticObjectId(institution_id))
        if existing:
            errors.append(f"Row {idx}: Batch '{name}' already exists")
            continue

        batch = Batch(
            name=name,
            institution_id=PydanticObjectId(institution_id),
        )
        await batch.insert()
        created += 1

    return api_response(200, "Batches uploaded", data={"created": created, "errors": errors})


@router.post("/testcases/{id}")
async def bulk_upload_testcases(
    id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(admin_only),
):
    problem_id = id

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    created = 0
    errors: list[str] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        tc_input = str(row[0]).strip()
        tc_output = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        tc_type = str(row[2]).strip().upper() if len(row) > 2 and row[2] else "HIDDEN"

        if tc_type not in (TestcaseType.EXAMPLE, TestcaseType.HIDDEN):
            tc_type = TestcaseType.HIDDEN

        tc = ProblemTestCase(
            input=tc_input,
            output=tc_output,
            type=tc_type,
            problem_id=PydanticObjectId(problem_id),
        )
        await tc.insert()
        created += 1

    return api_response(200, "Testcases uploaded", data={"created": created, "errors": errors})


@router.post("/cloud-info/")
async def bulk_upload_cloud_info(
    file: UploadFile = File(...),
    current_user: dict = Depends(admin_only),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    updated = 0
    errors: list[str] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        student_email = str(row[0]).strip()
        cloud_id = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        cloud_provider = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ""

        student = await Student.find_one(Student.email == student_email)
        if not student:
            errors.append(f"Row {idx}: Student {student_email} not found")
            continue

        student.cloud_id = cloud_id
        student.cloud_provider = cloud_provider
        student.updated_at = datetime.now(timezone.utc)
        await student.save()
        updated += 1

    return api_response(200, "Cloud info updated", data={"updated": updated, "errors": errors})


@router.post("/assignment/{id}")
async def bulk_upload_assignment(
    id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(not_student),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows_data.append({
            "question": str(row[0]).strip(),
            "options": [str(row[i]).strip() for i in range(1, 5) if len(row) > i and row[i]],
            "correct_option": str(row[5]).strip() if len(row) > 5 and row[5] else "",
        })

    return api_response(200, "Assignment data parsed", data={"rows": rows_data, "count": len(rows_data)})


@router.post("/assessment/{id}")
async def bulk_upload_assessment(
    id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(not_student),
):
    section_id = id

    section = await AssessmentSection.get(PydanticObjectId(section_id))
    if not section:
        return api_response(404, "Section not found", error="Not found")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        question = str(row[0]).strip()
        options = [str(row[i]).strip() for i in range(1, 5) if len(row) > i and row[i]]
        correct_option = str(row[5]).strip() if len(row) > 5 and row[5] else None
        max_marks = int(row[6]) if len(row) > 6 and row[6] else 1

        q = AssessmentQuestion(
            question=question,
            options=options,
            correct_option=correct_option,
            section_id=section.id,
            max_marks=max_marks,
        )
        await q.insert()
        created += 1

    return api_response(200, "Assessment questions uploaded", data={"created": created})
