import io
from datetime import datetime, timezone
from beanie import PydanticObjectId
from openpyxl import Workbook
from services.s3 import upload_file_to_s3

# These imports will work after Beanie is initialized in main.py
# We import models at module level since they're used in async functions after init

# Add parent python-server models path — worker shares same DB models
# Models are re-declared below for the worker context


async def process_assessment_report(data: dict):
    """Generate an XLSX report for an assessment and upload to S3."""
    from models import Assessment, AssessmentSubmission, AssessmentSection, AssessmentQuestion, AssessmentQuestionSubmission, Student

    assessment_id = data.get("assessment_id")
    if not assessment_id:
        print("No assessment_id in message")
        return

    assessment = await Assessment.get(PydanticObjectId(assessment_id))
    if not assessment:
        print(f"Assessment {assessment_id} not found")
        return

    print(f"Generating report for assessment: {assessment.name}")

    # Gather data
    submissions = await AssessmentSubmission.find(
        AssessmentSubmission.assessment_id == assessment.id
    ).to_list()

    sections = await AssessmentSection.find(
        AssessmentSection.assessment_id == assessment.id
    ).to_list()

    # Build XLSX
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Assessment Name", assessment.name])
    ws_summary.append(["Description", assessment.description or ""])
    ws_summary.append(["Start Time", str(assessment.start_time)])
    ws_summary.append(["End Time", str(assessment.end_time)])
    ws_summary.append(["Status", assessment.status])
    ws_summary.append(["Total Submissions", len(submissions)])

    # Submissions sheet
    ws_subs = wb.create_sheet("Submissions")
    ws_subs.append([
        "Student Name", "Student Email", "Total Marks",
        "Is Submitted", "Tab Switches", "Proctoring Status", "Submitted At"
    ])

    for sub in submissions:
        student = await Student.get(sub.student_id)
        ws_subs.append([
            student.name if student else "Unknown",
            student.email if student else "",
            sub.total_marks,
            "Yes" if sub.is_submitted else "No",
            sub.tab_switch_count,
            sub.proctoring_status,
            str(sub.submitted_at) if sub.submitted_at else "",
        ])

    # Per-section sheets with question-level detail
    for section in sections:
        ws_sec = wb.create_sheet(section.name[:31])  # Sheet name max 31 chars
        questions = await AssessmentQuestion.find(
            AssessmentQuestion.section_id == section.id
        ).to_list()

        header = ["Student Name", "Student Email"]
        for i, q in enumerate(questions, 1):
            header.append(f"Q{i} Answer")
            header.append(f"Q{i} Marks")
        header.append("Section Total")
        ws_sec.append(header)

        for sub in submissions:
            student = await Student.get(sub.student_id)
            row = [
                student.name if student else "Unknown",
                student.email if student else "",
            ]
            section_total = 0
            for q in questions:
                q_sub = await AssessmentQuestionSubmission.find_one(
                    AssessmentQuestionSubmission.question_id == q.id,
                    AssessmentQuestionSubmission.student_id == sub.student_id,
                    AssessmentQuestionSubmission.assessment_id == assessment.id,
                )
                row.append(q_sub.answer if q_sub else "")
                marks = q_sub.marks_obtained if q_sub else 0
                row.append(marks)
                section_total += marks
            row.append(section_total)
            ws_sec.append(row)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Upload to S3
    filename = f"assessment-report-{assessment_id}.xlsx"
    url = upload_file_to_s3(
        buffer.read(),
        "assessment-reports",
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Update assessment with report URL
    assessment.report = url
    assessment.report_status = "PROCESSED"
    assessment.updated_at = datetime.now(timezone.utc)
    await assessment.save()

    print(f"Report uploaded: {url}")
